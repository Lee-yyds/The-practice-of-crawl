import requests
from lxml import etree
from openpyxl import Workbook
from multiprocessing import Pool


def get_data(car_id, page):
    headers = {
        "authority": "price.pcauto.com.cn",
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "cache-control": "no-cache",
        "pragma": "no-cache",
        "referer": "https://price.pcauto.com.cn/comment/sg23653/p2.html",
        "sec-ch-ua": "^\\^Microsoft",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "^\\^Windows^^",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.42"
    }
    cookies = {
        "pcLocate": "^%^7B^%^22proCode^%^22^%^3A^%^22510000^%^22^%^2C^%^22pro^%^22^%^3A^%^22^%^E5^%^9B^%^9B^%^E5^%^B7^%^9D^%^22^%^2C^%^22proId^%^22^%^3A12^%^2C^%^22cityCode^%^22^%^3A^%^22510500^%^22^%^2C^%^22city^%^22^%^3A^%^22^%^E6^%^B3^%^B8^%^E5^%^B7^%^9E^%^22^%^2C^%^22cityId^%^22^%^3A425^%^2C^%^22url^%^22^%^3A^%^22^%^2F^%^2Fwww.pcauto.com.cn^%^2Fqcbj^%^2Fsc^%^2Fluzhou^%^2F^%^22^%^2C^%^22dataType^%^22^%^3A^%^22geolocation^%^22^%^2C^%^22version^%^22^%^3A^%^222.0.0^%^22^%^2C^%^22msg^%^22^%^3A^%^220^%^22^%^2C^%^22point^%^22^%^3A^%^7B^%^22lng^%^22^%^3A105.51330409194^%^2C^%^22lat^%^22^%^3A28.197799870937^%^7D^%^2C^%^22expires^%^22^%^3A1665112941881^%^7D",
        "AplocationWap": "^%^7B^%^22regionId^%^22^%^3A425^%^2C^%^22regionName^%^22^%^3A^%^22^%^E6^%^B3^%^B8^%^E5^%^B7^%^9E^%^22^%^7D",
        "iwt_uuid": "7c77b504-1c89-412a-a021-0bfe1369429c"
    }
    url = f"https://price.pcauto.com.cn/comment/sg{car_id}/p{page}.html"
    print(url)
    response = requests.get(url, headers=headers, cookies=cookies)
    # print(response)
    return response


def parse_data(response, ws1, row):
    tree = etree.HTML(response.text)
    for i in range(1, 11):
        li = tree.xpath(f'//div[@class="litDy clearfix"][{i}]//div[@class="fzbox"]/ul/li/b/text()')
        # title = tree.xpath(f'//div[@class="litDy clearfix"][{i}]//div[@class="fzbox"]/ul/li/span/text()')
        # title1 = tree.xpath(f'//div[@class="litDy clearfix"][{i}]//div[@class="dianPing clearfix"]/div/b/text()')
        li1 = tree.xpath(f'//div[@class="litDy clearfix"][{i}]//div[@class="dianPing clearfix"]/div/span/text()')
        print(li1)
        if li:
            # print(title)
            # print(li1)
            # print(content)
            ws1.append(li)
            for n in range(len(li1)):
                # ws1.cell(row=row, column=10 + n).value = title1[n]
                ws1.cell(row=row, column=10+n).value = li1[n]
            row += 1
        else:
            break


def run(car_id):
    wb = Workbook()
    ws1 = wb.active
    ws1.append(['外观', '内饰', '空间', '配置', '动力', '越野', '油耗', '舒适', '', '优点', '缺点', '外观', '内饰', '空间', '配置', '动力', '越野', '油耗', '舒适'])
    row = 2
    for i in range(1, 261):
        t = get_data(car_id, i)
        parse_data(t, ws1, row)
        row += 10
    wb.save('蔚来ES6.xlsx')


if __name__ == '__main__':
    pool = Pool(processes=6)
    pool.map(run, (4881,))
    pool.close()
    pool.join()
    # run(23323)
